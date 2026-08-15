"""Document Service：所有文件只讀同一份 ReportPayload，不各自重查資料。

輸出：
- report_payload.json（內部除錯／前端預覽）
- design_engineering_proposal.html（對外提案書）
- estimate_and_schedule.xlsx（對外只有「工程估價」「初步排程」兩張工作表）
"""
from __future__ import annotations

import json
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..adapters.ports import ProjectRepository
from ..contracts import DocumentManifest, ReportPayload


class DocumentService:
    def __init__(
        self,
        generated_dir: Path,
        template_dir: Path,
        repository: ProjectRepository,
        api_prefix: str,
    ) -> None:
        self.generated_dir = generated_dir
        self.repository = repository
        self.api_prefix = api_prefix.rstrip("/")
        self.template_dir = template_dir

    def _environment(self):
        # ponytail: 只有 report_html 要 jinja2；XLSX/JSON 不該被這個 import 綁架
        from jinja2 import Environment, FileSystemLoader, select_autoescape

        return Environment(
            loader=FileSystemLoader(str(self.template_dir)),
            autoescape=select_autoescape(["html", "xml"]),
        )

    def render(
        self, report: ReportPayload, requested_types: list[str]
    ) -> ReportPayload:
        package_dir = (
            self.generated_dir / report.project_id / report.revision / report.package_id
        )
        package_dir.mkdir(parents=True, exist_ok=True)

        supported = {
            "report_json": ("report_payload.json", "application/json"),
            "report_html": (
                "design_engineering_proposal.html",
                "text/html; charset=utf-8",
            ),
            "estimate_xlsx": (
                "estimate_and_schedule.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        }
        manifests: list[DocumentManifest] = []
        paths: dict[str, Path] = {}

        for document_type in requested_types:
            if document_type not in supported:
                continue
            file_name, content_type = supported[document_type]
            document_id = f"doc_{uuid4().hex[:12]}"
            manifest = DocumentManifest(
                document_id=document_id,
                document_type=document_type,
                file_name=file_name,
                content_type=content_type,
                download_url=f"{self.api_prefix}/documents/{document_id}/download",
            )
            manifests.append(manifest)
            paths[document_type] = package_dir / file_name

        final_report = report.model_copy(update={"documents": manifests}, deep=True)

        if "report_json" in paths:
            paths["report_json"].write_text(
                json.dumps(
                    final_report.model_dump(mode="json"),
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )

        if "report_html" in paths:
            template = self._environment().get_template("report.html.j2")
            html = template.render(report=final_report.model_dump(mode="json"))
            paths["report_html"].write_text(html, encoding="utf-8")

        if "estimate_xlsx" in paths:
            self._write_xlsx(final_report, paths["estimate_xlsx"])

        for manifest in manifests:
            self.repository.register_document(manifest, paths[manifest.document_type])

        return final_report

    @staticmethod
    def _write_xlsx(report: ReportPayload, path: Path) -> None:
        """MVP 對外 Excel 僅保留兩張工作表：工程估價與初步排程。"""
        workbook = Workbook()
        estimate = workbook.active
        estimate.title = "工程估價"

        estimate.append(["RoomPilot 初步工程估價", None, None, None])
        estimate.append(["專案ID", report.project_id, "版本", report.revision])
        estimate.append(
            ["產生時間", report.generated_at.isoformat(), "幣別", report.estimate.currency]
        )
        estimate.append(
            [
                "已知小計",
                report.estimate.known_subtotal,
                "待詢價工項",
                report.estimate.pending_quote_count,
            ]
        )
        estimate.append(["重要聲明", report.estimate.disclaimer, None, None])
        estimate.append([])
        estimate.append([
            "房間ID", "工種", "工項編碼", "工項名稱", "原始量", "損耗率",
            "計價量", "單位", "材料單價", "人工單價", "其他單價", "小計",
            "狀態", "價格地區", "價格日期", "來源", "信心",
        ])
        for line in report.estimate.lines:
            estimate.append([
                line.room_id, line.trade, line.work_item_code, line.name,
                line.raw_quantity, line.waste_rate, line.pricing_quantity, line.unit,
                line.material_unit_price, line.labor_unit_price, line.other_unit_price,
                line.subtotal if line.subtotal is not None else "待詢價",
                line.status, line.price_region,
                line.price_effective_date.isoformat()
                if line.price_effective_date
                else None,
                line.price_source, line.confidence,
            ])

        schedule = workbook.create_sheet("初步排程")
        schedule.append(["RoomPilot 初步工程排程", None, None, None])
        schedule.append(["專案ID", report.project_id, "版本", report.revision])
        schedule.append(
            [
                "預估總工期(日)",
                report.schedule.estimated_total_days,
                "資料不足項目",
                report.schedule.unknown_duration_count,
            ]
        )
        schedule.append(["重要聲明", report.schedule.disclaimer, None, None])
        schedule.append([])
        schedule.append([
            "Task ID", "房間ID", "工項編碼", "工作", "工程量", "單位",
            "日產能", "工班", "準備日", "施工日", "等待日", "總工日",
            "開始日", "完成日", "前置 Task", "狀態", "來源", "信心",
        ])
        for task in report.schedule.tasks:
            schedule.append([
                task.task_id, task.room_id, task.work_item_code, task.name,
                task.quantity, task.unit, task.daily_productivity, task.crew_count,
                task.preparation_days, task.construction_days, task.waiting_days,
                task.total_days, task.start_day, task.finish_day,
                ",".join(task.predecessor_task_ids), task.status, task.source,
                task.confidence,
            ])

        title_fill = PatternFill("solid", fgColor="16324F")
        header_fill = PatternFill("solid", fgColor="2F6B8A")
        white_font = Font(color="FFFFFF", bold=True)

        for sheet, header_row in ((estimate, 7), (schedule, 6)):
            sheet.freeze_panes = f"A{header_row + 1}"
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
            sheet.cell(1, 1).fill = title_fill
            sheet.cell(1, 1).font = Font(color="FFFFFF", bold=True, size=14)
            for cell in sheet[header_row]:
                cell.fill = header_fill
                cell.font = white_font
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            sheet.auto_filter.ref = (
                f"A{header_row}:{sheet.cell(header_row, sheet.max_column).coordinate}"
            )
            for column_index in range(1, sheet.max_column + 1):
                max_length = max(
                    len(str(sheet.cell(row_index, column_index).value))
                    if sheet.cell(row_index, column_index).value is not None
                    else 0
                    for row_index in range(1, sheet.max_row + 1)
                )
                letter = get_column_letter(column_index)
                sheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 42)
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    # 房間 id 與材料名稱是前端寫的。openpyxl 看到 "=" 開頭就存成公式，
                    # 而這份檔是要寄給屋主與工班的，不能在對方的 Excel 裡執行。
                    if cell.data_type == "f":
                        cell.data_type = "s"

        workbook.save(path)
