"""交付提案 PDF（roompilot-delivery-pdf 打包 skill）adapter 與 FastAPI 測試。

content.json 底稿與 LLM 合稿不需排版引擎，離線可測；實際 Chromium 排版與
端到端下載在 playwright 可用時執行（未安裝則 skip，API 必須明確 503）。
"""
from __future__ import annotations

import base64
import importlib.util
import io
import json
import re
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from PIL import Image

from backend.agent.skills.delivery import DeliverySkill, build_content
from backend.server import design_manual_service, main
from backend.server.design_manual_service import (
    _assemble_store,
    create_delivery_proposal,
)
from backend.server.project_store import ProjectStore

PLAYWRIGHT_AVAILABLE = importlib.util.find_spec("playwright") is not None
OPENPYXL_AVAILABLE = importlib.util.find_spec("openpyxl") is not None


def _png_b64(color=(200, 180, 150)) -> str:
    buffer = io.BytesIO()
    Image.new("RGB", (64, 64), color).save(buffer, format="PNG")
    return base64.b64encode(buffer.getvalue()).decode("ascii")


def _scene() -> dict:
    return {
        "scene_id": "scene-x",
        "requirement": {
            "style": "scandinavian",
            "constraints": {"notes": ["保留閱讀角落"]},
        },
        "style_card": {
            "card_id": "scandinavian_1",
            "name_zh": "自然木質",
            "palette_hex": ["#F3EBDD", "#D3B48A"],
        },
        "design_choices": {"floor_option": "wood_oak", "wall_option": "auto"},
        "surface_catalog": {
            "surfaces": [
                {"surface_id": "wood_oak", "name_zh": "橡木地板", "color_zh": "暖木色"}
            ]
        },
        "floorplan": {"width_cm": 400, "depth_cm": 360},
        "scene_objects": [
            {
                "id": "sofa-1",
                "furniture_id": "sofa-1",
                "instance_id": "sofa-1#1",
                "normalized_type": "sofa",
                "name_zh_raw": "北歐布沙發",
                "material": "亞麻布",
                "primary_style": "scandinavian",
                "price_twd": 18800,
                "size_cm": {"width": 210, "depth": 90, "height": 85},
                "placement_room_id": "living-1",
            }
        ],
    }


def _rooms(with_image: bool = True) -> list[dict]:
    return [
        {
            "room_id": "living-1",
            "room_label": "客廳",
            "width_cm": 400,
            "depth_cm": 360,
            "image_data_url": (
                f"data:image/png;base64,{_png_b64()}" if with_image else None
            ),
        },
        {"room_id": "study-1", "room_label": "書房", "width_cm": 300, "depth_cm": 280},
        # 全部房型都排篇章（含浴室、陽台）。
        {"room_id": "bath-1", "room_label": "浴室", "width_cm": 200, "depth_cm": 180},
        {"room_id": "balcony-1", "room_label": "陽台", "width_cm": 300, "depth_cm": 120},
    ]


def _docs():
    """經 server 組裝取得 agent 文件（與正式路徑同一條組裝）。"""
    store, _ = _assemble_store(_scene(), _rooms(), design_revision=3)
    from backend.agent.documents import DocKey, LayoutDoc, RequirementDoc, SceneDoc

    snapshot = store.snapshot()
    return (
        RequirementDoc.from_dict(snapshot[DocKey.REQUIREMENTS]),
        LayoutDoc.from_dict(snapshot[DocKey.LAYOUT]),
        SceneDoc.from_dict(snapshot[DocKey.variant(DocKey.SCENE, "chosen")]),
    )


# ------------------------------------------------------------ content 底稿


def test_offline_content_is_factual_and_clean() -> None:
    requirements, layout, scene = _docs()
    content = build_content(
        "林宅", requirements, layout, scene,
        {"living-1": "rooms/living-1.png"}, design_revision=3,
    )

    assert content["meta"]["project_name"] == "林宅"
    assert content["meta"]["cover_image"] == "rooms/living-1.png"
    assert "revision 3" in content["meta"]["version"]

    rooms = {room["name"]: room for room in content["rooms"]}
    # 全部房型都排篇章（含浴室、陽台）。
    assert set(rooms) == {"客廳", "書房", "浴室", "陽台"}
    living = rooms["客廳"]
    assert len(living["look"]) >= 40
    # 選件依據一條就夠；擺位與淨空是工作紀錄，不是屋主要讀的設計理由。
    assert living["rationale"][0]["title"] == "選件依據"
    assert not any(
        word in json.dumps(living["rationale"], ensure_ascii=False)
        for word in ("擺位", "淨空", "幾何引擎", "不重疊")
    )
    spec_text = json.dumps(living["specs"], ensure_ascii=False)
    assert "北歐布沙發" in spec_text and "210×90 cm" in spec_text
    assert "18,800" not in spec_text, "金額只出現在報價單章節，規格表不帶價"
    # 型號留在規格表，敘述只用通用中文名。
    assert "北歐布沙發" in living["look"] and "sofa" not in living["look"]

    # 設計總論＝後續每一章的摘要（章名當標題）。
    titles = [pillar["title"] for pillar in content["statement"]["pillars"]]
    assert titles[0] == "全案速覽" and titles[-1] == "接下來"
    assert "客廳" in titles and "書房" in titles and "色彩與材質" in titles

    # 全案速覽不能只有幾格數字：要有開場與空間一覽表。
    assert content["overview"]["intro"]
    assert len(content["overview"]["table"]["rows"]) == 4  # 全部房型都列

    # 無圖房間必須寫進 appendix.limits（不能讓屋主以為漏做）。
    assert any("書房" in limit for limit in content["appendix"]["limits"])
    # 未標價原則保留。
    assert any("正式報價" in limit for limit in content["appendix"]["limits"])
    # 60/30/10 色卡 swatches：色名要看得懂，不是把色碼印兩次。
    swatch = content["palette"]["swatches"][0]
    assert swatch["usage"].startswith("主色")
    assert swatch["name"].startswith("米白") and swatch["hex"] in swatch["name"]
    assert content["palette"]["intro"].count("\n") >= 2
    assert all(material["why"] for material in content["materials"])

    # deterministic 底稿不得出現廣告腔／AI 高頻詞（writing-rules 禁詞抽查）。
    serialized = json.dumps(content, ensure_ascii=False)
    for banned in ("打造", "極致", "匠心", "營造出", "坐落於", "此外", "藉由", "不只是"):
        assert banned not in serialized, banned


def test_living_room_night_image_becomes_an_extra_image() -> None:
    """客廳夜間圖走 `extra_images`，不搶主視覺、不進封面。

    schema 早就支援 `rooms[].extra_images`（content-schema.md 也寫明只給客廳、
    主臥這種重點空間），先前只是 `_write_room_images()` 一房只落一張圖，
    夜間圖根本沒有被寫進 content.json。
    """
    requirements, layout, scene = _docs()
    content = build_content(
        "林宅", requirements, layout, scene,
        {"living-1": "rooms/living-1.png"},
        night_image_files={"living-1": "rooms/living-1_night.png"},
        design_revision=3,
    )

    rooms = {room["name"]: room for room in content["rooms"]}
    living = rooms["客廳"]
    assert living["hero_image"] == "rooms/living-1.png"     # 主視覺仍是日光
    # 兩張並列時左邊那張要講清楚是日光，只寫「最終渲染」屋主分不出差別。
    assert living["hero_caption"] == "客廳日光。"
    assert living["extra_images"] == [
        {
            "src": "rooms/living-1_night.png",
            "caption": "客廳夜間燈光。同一視角、同一色卡，只換光影。",
        }
    ]
    # 封面用日光那張；夜間圖仍要進檔案清單，屋主才知道收到幾張圖。
    assert content["meta"]["cover_image"] == "rooms/living-1.png"
    names = {row["name"] for row in content["appendix"]["files"]}
    assert {"rooms/living-1.png", "rooms/living-1_night.png"} <= names
    # 沒有夜間圖的空間不長出空的 extra_images，標題也維持原樣。
    assert "extra_images" not in rooms["書房"]
    assert rooms["浴室"]["hero_caption"] == ""      # 無圖就沒有圖說


def _load_build_pdf():
    """打包 skill 的資料夾名有連字號，不能 import，只能按路徑載入。"""
    path = (
        Path(__file__).resolve().parents[1]
        / "backend/agent/skills/roompilot-delivery-pdf/scripts/build_pdf.py"
    )
    spec = importlib.util.spec_from_file_location("rp_build_pdf", path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def test_room_hero_and_extra_images_render_side_by_side(tmp_path) -> None:
    """客廳篇章的兩張圖並列：左邊日光（hero）、右邊夜間，同一個容器內。

    夜間圖排在頁尾要翻著比，而且會把客廳推到第二頁。
    """
    build_pdf = _load_build_pdf()
    rooms_dir = tmp_path / "rooms"
    rooms_dir.mkdir()
    for name, color in (("day.png", (200, 180, 150)), ("night.png", (20, 24, 40))):
        (rooms_dir / name).write_bytes(base64.b64decode(_png_b64(color)))

    html = build_pdf.render_room(
        "03",
        {
            "name": "客廳",
            "hero_image": "rooms/day.png",
            "hero_caption": "客廳日光。",
            "extra_images": [{"src": "rooms/night.png", "caption": "客廳夜間燈光。"}],
        },
        tmp_path,
    )

    grid = re.search(r'<div class="img-grid hero">(.*?)</div>', html, re.S)
    assert grid, "有附圖時主視覺要進並列容器"
    pair = grid.group(1)
    assert pair.count("<img") == 2 and html.count("<img") == 2
    assert pair.index("客廳日光。") < pair.index("客廳夜間燈光。"), "日光在左、夜間在右"

    # 只有一張圖的空間維持滿版主視覺，不要為了一張圖開兩欄。
    solo = build_pdf.render_room(
        "04", {"name": "書房", "hero_image": "rooms/day.png"}, tmp_path
    )
    assert "img-grid" not in solo and '<figure class="hero">' in solo


def test_write_room_images_lands_day_and_night_as_separate_files(tmp_path) -> None:
    """夜間圖要另外落一個檔名；同名會直接蓋掉日光那張。"""
    from backend.agent.documents import (
        ImageLibraryDoc,
        ImageRecord,
        LayoutDoc,
        LayoutRoom,
    )

    images = ImageLibraryDoc()
    for image_id, room_id, stage, color in (
        ("img_day", "living-1", "full_render", (200, 180, 150)),
        ("img_night", "living-1", "full_render_night", (20, 24, 40)),
        ("img_study", "study-1", "full_render", (200, 180, 150)),
    ):
        images.records.append(
            ImageRecord(
                image_id=image_id,
                room_id=room_id,
                stage=stage,
                model="google/gemini-3.1-flash-image",
                image_ref=_png_b64(color),
                seq=images.next_seq(),
            )
        )
    layout = LayoutDoc(
        rooms=[
            LayoutRoom(room_id="living-1", name="客廳", width_cm=400, depth_cm=360),
            LayoutRoom(room_id="study-1", name="書房", width_cm=300, depth_cm=260),
        ],
        source="scene_json",
    )

    day, night = DeliverySkill(None)._write_room_images(images, layout, tmp_path)

    assert day == {"living-1": "rooms/living-1.png", "study-1": "rooms/study-1.png"}
    assert night == {"living-1": "rooms/living-1_night.png"}   # 只有客廳有夜間圖
    rooms_dir = tmp_path / "rooms"
    assert (rooms_dir / "living-1_night.png").read_bytes() != (
        rooms_dir / "living-1.png"
    ).read_bytes()


def test_money_lives_only_in_the_quote_chapter() -> None:
    """設計章節不談錢：金額集中在報價單，缺價的品項標「待報價」不補猜。"""
    requirements, layout, scene = _docs()
    content = build_content(
        "林宅", requirements, layout, scene, {}, design_revision=3,
    )

    quote = content["quote"]
    assert quote["table"]["columns"] == ["空間", "品項", "規格", "數量", "單價", "小計"]
    sofa = next(row for row in quote["table"]["rows"] if "沙發" in row[1])
    assert sofa[0] == "客廳" and sofa[4] == "18,800 元" and sofa[5] == "18,800 元"
    assert any(fact["label"] == "已標價合計" for fact in quote["summary"])

    # 除了報價單，其他區塊一毛錢都不能出現。
    without_quote = {k: v for k, v in content.items() if k != "quote"}
    assert not re.search(
        r"\d[\d,]*\s*元", json.dumps(without_quote, ensure_ascii=False)
    )


def test_identical_furniture_collapses_into_one_spec_row() -> None:
    """四張一樣的餐椅抄四遍，屋主第一反應是「這表格是不是壞了」。"""
    scene = _scene()
    chair = {
        "id": "chair",
        "normalized_type": "dining-chair",
        "name_zh_raw": "鉚釘現代餐椅，34 英寸（約 86.4 釐米）高, 粉筆色",
        "material": "brass",
        "size_cm": {"width": 51, "depth": 52, "height": 86},
        "placement_room_id": "living-1",
    }
    scene["scene_objects"] += [
        {**chair, "instance_id": f"chair#{index}"} for index in range(1, 4)
    ]
    store, _ = _assemble_store(scene, _rooms(), design_revision=1)
    from backend.agent.documents import DocKey, LayoutDoc, RequirementDoc, SceneDoc

    snapshot = store.snapshot()
    content = build_content(
        "林宅",
        RequirementDoc.from_dict(snapshot[DocKey.REQUIREMENTS]),
        LayoutDoc.from_dict(snapshot[DocKey.LAYOUT]),
        SceneDoc.from_dict(snapshot[DocKey.variant(DocKey.SCENE, "chosen")]),
        {},
    )
    living = next(room for room in content["rooms"] if room["name"] == "客廳")
    chairs = [row for row in living["specs"] if "餐椅" in row["label"]]
    assert len(chairs) == 1
    assert "共 3 件" in chairs[0]["value"]
    # 類型欄出中文，不是 dining-chair。
    assert "餐椅，51×52 cm" in chairs[0]["value"] and "brass" not in chairs[0]["value"]
    # 敘述用通用名，不重複列三次。
    assert living["look"].count("餐椅") == 1


def test_llm_copy_merges_into_content() -> None:
    requirements, layout, scene = _docs()
    content = build_content("林宅", requirements, layout, scene, {}, design_revision=1)

    class CopyGateway:
        available = True

        def chat(self, messages, *, model=None, temperature=0.3, force_json=False, reasoning=None):
            return json.dumps(
                {
                    "statement": {
                        "hook": "從玄關進門會先看到整面南向的光。",
                        "pillars": [
                            {"title": "把光留下來", "body": "南向開窗前不放高櫃。"},
                            {"title": "收納藏起來", "body": "櫃體全部靠牆做滿。"},
                        ],
                    },
                    "rooms": [
                        {
                            "room_id": "living-1",
                            "scene_line": "週五晚上四個人擠在沙發上看片。",
                            "look": "沙發背對走道，回家坐下時不會有人從背後經過；"
                                    "布面是亞麻的，夏天不黏腿。",
                            "rationale": [
                                {"title": "沙發背對走道", "body": "你說常有朋友來，動線留在沙發後面。"},
                                {"title": "留白不塞櫃", "body": "多一個櫃走道會縮，兩人錯身會卡。"},
                            ],
                        }
                    ],
                },
                ensure_ascii=False,
            )

    merged = DeliverySkill(CopyGateway())._merge_llm_copy(
        content, requirements, layout, scene
    )

    assert merged is True
    living = next(room for room in content["rooms"] if room["room_id"] == "living-1")
    assert living["scene_line"] == "週五晚上四個人擠在沙發上看片。"
    assert "亞麻" in living["look"]
    assert living["rationale"][0]["title"] == "沙發背對走道"
    assert content["statement"]["hook"] == "從玄關進門會先看到整面南向的光。"
    # 書房未回傳 → 保留 deterministic 底稿。
    study = next(room for room in content["rooms"] if room["room_id"] == "study-1")
    assert "書房" in study["look"]


def test_llm_placement_rationale_is_dropped() -> None:
    """LLM 若又寫回「擺位與淨空」，合稿時擋掉——那是工作紀錄不是設計理由。"""
    requirements, layout, scene = _docs()
    content = build_content("林宅", requirements, layout, scene, {}, design_revision=1)

    class NoisyGateway:
        available = True

        def chat(self, messages, *, model=None, temperature=0.3, force_json=False, reasoning=None):
            return json.dumps(
                {
                    "overview_intro": "28.6 坪、兩個空間，材質與配色同一套。",
                    "palette_intro": "米白鋪底，木色收邊，點綴只出現在小物件上。",
                    "rooms": [
                        {
                            "room_id": "living-1",
                            "look": "從玄關轉進來，第一眼是整面南向的窗。",
                            "rationale": [
                                {"title": "擺位與淨空", "body": "位置由幾何引擎驗證，不重疊。"},
                                {"title": "選件依據", "body": "你說很少看電視，電視收在側牆淺櫃。"},
                            ],
                        }
                    ],
                },
                ensure_ascii=False,
            )

    DeliverySkill(NoisyGateway())._merge_llm_copy(content, requirements, layout, scene)

    living = next(room for room in content["rooms"] if room["room_id"] == "living-1")
    assert [row["title"] for row in living["rationale"]] == ["選件依據"]
    assert content["overview"]["intro"].startswith("28.6 坪")
    assert content["palette"]["intro"].startswith("米白鋪底")


def test_offline_copy_is_reported_not_silently_downgraded() -> None:
    """沒有 LLM 照樣出檔，但不能假裝文案是寫過的。"""
    requirements, layout, scene = _docs()
    content = build_content("林宅", requirements, layout, scene, {}, design_revision=1)
    assert DeliverySkill(None)._merge_llm_copy(content, requirements, layout, scene) is False


# ------------------------------------------------------------ PDF 端到端


@pytest.mark.skipif(not PLAYWRIGHT_AVAILABLE, reason="playwright 未安裝")
def test_delivery_pdf_end_to_end(tmp_path) -> None:
    result, record = create_delivery_proposal(
        "proj12345678", "林宅", _scene(), _rooms(), tmp_path,
        design_revision=3, gateway=object(),
    )
    pdf = (tmp_path / record["filename"]).read_bytes()
    assert pdf[:4] == b"%PDF"
    assert record["filename"].startswith("roompilot-proposal-proj1234-")
    assert record["rendered_rooms"] == ["living-1"]
    # content.json 留存於 PDF 旁供數字溯源。
    content = json.loads(
        (tmp_path / record["filename"]).with_suffix(".content.json").read_text("utf-8")
    )
    assert content["meta"]["project_name"] == "林宅"

    import pikepdf

    with pikepdf.open(tmp_path / record["filename"]) as doc:
        # 封面＋總論＋速覽＋兩房＋色彩材質＋接下來 ≈ 7 頁，至少要有 5 頁。
        assert len(doc.pages) >= 5


def _client(tmp_path, monkeypatch) -> tuple[TestClient, str]:
    monkeypatch.delenv("OPENROUTER_API_KEY", raising=False)
    monkeypatch.setattr(main, "PROJECT_STORE", ProjectStore(tmp_path / "runtime"))
    client = TestClient(main.app)
    project = client.post("/api/projects", json={"name": "交付提案"}).json()["project"]
    return client, project["project_id"]


@pytest.mark.skipif(not PLAYWRIGHT_AVAILABLE, reason="playwright 未安裝")
def test_generate_then_download_proposal(tmp_path, monkeypatch) -> None:
    client, project_id = _client(tmp_path, monkeypatch)
    created = client.post(
        f"/api/projects/{project_id}/delivery-proposal",
        json={"project_id": project_id, "scene": _scene(), "rooms": _rooms()},
    )
    assert created.status_code == 201
    body = created.json()
    assert body["proposal"]["download_url"] == (
        f"/api/projects/{project_id}/delivery-proposal/pdf"
    )
    assert "filename" not in body["proposal"]

    downloaded = client.get(f"/api/projects/{project_id}/delivery-proposal/pdf")
    assert downloaded.status_code == 200
    assert downloaded.headers["content-type"] == "application/pdf"
    assert downloaded.content[:4] == b"%PDF"

    project = client.get(f"/api/projects/{project_id}").json()["project"]
    assert project["workflow"]["delivery_proposal"]["filename"].endswith(".pdf")


def test_engine_missing_reports_503(tmp_path, monkeypatch) -> None:
    client, project_id = _client(tmp_path, monkeypatch)
    monkeypatch.setattr(
        design_manual_service,
        "delivery_engine_status",
        lambda: (False, "尚未安裝交付提案排版引擎（測試）。"),
    )
    status = client.get("/api/delivery-proposal/status")
    assert status.json() == {
        "available": False,
        "reason": "尚未安裝交付提案排版引擎（測試）。",
    }
    created = client.post(
        f"/api/projects/{project_id}/delivery-proposal",
        json={"project_id": project_id, "scene": _scene(), "rooms": _rooms()},
    )
    assert created.status_code == 503
    assert created.json()["detail"]["code"] == "delivery_engine_not_configured"


def test_download_before_generation_is_404(tmp_path, monkeypatch) -> None:
    client, project_id = _client(tmp_path, monkeypatch)
    response = client.get(f"/api/projects/{project_id}/delivery-proposal/pdf")
    assert response.status_code == 404
    assert response.json()["detail"]["code"] == "delivery_proposal_not_found"


# 第 8 步「產出設計提案」同一顆按鈕要吐兩份檔：排版 PDF 與工程估價 XLSX。
# 排版引擎換成假的，這裡只驗第二份檔的接線（估價 → workflow 紀錄 → 下載）。
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl 未安裝")
def test_delivery_proposal_also_writes_the_engineering_estimate(
    tmp_path, monkeypatch
) -> None:
    client, project_id = _client(tmp_path, monkeypatch)
    assert client.get(f"/api/projects/{project_id}/delivery-proposal/xlsx").json()[
        "detail"
    ]["code"] == "engineering_estimate_not_found"

    monkeypatch.setattr(
        main,
        "create_delivery_proposal",
        lambda *args, **kwargs: (None, {"filename": "fake.pdf", "warnings": []}),
    )
    main.PROJECT_STORE.update_workflow(
        project_id,
        workflow={
            "space_confirmation": {
                "coordinate_unit": "cm",
                "schema_version": "2.0",
                "rooms": [
                    {
                        "id": "room-1",
                        "label": "客廳",
                        "type": "living_room",
                        "confirmed": True,
                        "polygon_cm": [
                            {"x": 0, "y": 0},
                            {"x": 400, "y": 0},
                            {"x": 400, "y": 300},
                            {"x": 0, "y": 300},
                        ],
                    }
                ],
                "structures": {
                    "walls": [],
                    "doors": [],
                    "windows": [],
                    "beams": [],
                    "columns": [],
                },
            },
            # 問卷存的是型錄 ID，不是「乳膠漆」；牆與天花靠 knowledge 的
            # "paint" 關鍵字才對得上工項，拿掉就整批漏算。
            "requirements": {
                "finishes": {
                    "floorMaterial": "wood_tile_ccity_wood_look_tiles_cvt212022",
                    "wallMaterial": "wall_json_ambientcg_wall_paint_concrete036",
                    "ceilingMaterial": "flat-paint",
                }
            },
        },
    )

    created = client.post(
        f"/api/projects/{project_id}/delivery-proposal",
        json={"project_id": project_id, "scene": _scene(), "rooms": _rooms()},
    )
    assert created.status_code == 201
    estimate = created.json()["proposal"]["engineering"]
    assert estimate["status"] != "skipped", estimate.get("reason")
    assert estimate["line_count"] > 0
    assert {"PAINT-WALL", "PAINT-CEILING"} <= _work_item_codes(
        main.PROJECT_STORE.runtime_dir / "manuals" / estimate["file"]
    )

    downloaded = client.get(f"/api/projects/{project_id}/delivery-proposal/xlsx")
    assert downloaded.status_code == 200
    assert downloaded.content[:2] == b"PK"


# 對不到工項的材料曾經整批從估價單消失，封面還印「待詢價 0」加一個看似完整的總價。
# 漏算必須看得見：明細有一列、pending_quote 數不為零、總價變成 None。
@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl 未安裝")
def test_material_without_a_work_item_mapping_is_listed_not_dropped(tmp_path) -> None:
    from backend.server.engineering_report import build_engineering_estimate

    workflow = {
        "space_confirmation": {
            "coordinate_unit": "cm",
            "schema_version": "2.0",
            "rooms": [
                {
                    "id": "room-1",
                    "label": "客廳",
                    "type": "living_room",
                    "confirmed": True,
                    "polygon_cm": [
                        {"x": 0, "y": 0},
                        {"x": 400, "y": 0},
                        {"x": 400, "y": 300},
                        {"x": 0, "y": 300},
                    ],
                }
            ],
            "structures": {
                "walls": [],
                "doors": [],
                "windows": [],
                "beams": [],
                "columns": [],
            },
        },
        # 天花清水模在 knowledge 裡沒有對照工項，也不在面材型錄內。
        "requirements": {"finishes": {"ceilingMaterial": "exposed-concrete"}},
    }
    record = build_engineering_estimate("proj-unmapped", "1", workflow, tmp_path)
    assert record["status"] != "skipped", record.get("reason")
    assert record["estimated_total"] is None
    assert "UNMAPPED-CEILING" in _work_item_codes(tmp_path / record["file"])


def _work_item_codes(xlsx_path: Path) -> set[str]:
    from openpyxl import load_workbook

    sheet = load_workbook(xlsx_path)["工程估價"]
    return {
        str(row[2]) for row in sheet.iter_rows(min_row=8, values_only=True) if row[2]
    }


def _design_delivery_payload(project_id: str) -> dict:
    return {
        "project_id": project_id,
        "style_card": {
            "id": "scandinavian_1",
            "name": "自然木質",
            "palette_hex": ["#F3EBDD", "#D3B48A"],
            "email": "should-not-leak@example.com",
        },
        "configuration_snapshot": {
            "snapshot_id": "snap-1",
            "furniture": [
                {
                    "instance_id": "sofa-1#1",
                    "room_id": "living",
                    "name_zh": "北歐布沙發",
                    "price_twd": 12900,
                },
                {"instance_id": "table-1#1", "room_id": "living", "name_zh": "小茶几"},
            ],
            "fixed_structure": {"walls": [{}, {}], "doors": [{}], "windows": []},
        },
        "rooms": [
            {
                "room_id": "living",
                "room_name": "客廳",
                "room_type": "living_room",
                "questionnaire": {
                    "summary": "用途：家庭聚會；已選家具：北歐布沙發",
                    "usage": ["家庭聚會"],
                    "lockedFurniture": ["北歐布沙發"],
                    "surfaces": {"wallDefault": "warm_white", "floor": "light_oak"},
                    "generativeEquipment": {"primaryUse": "影音娛樂"},
                },
                "view": {"camera": {"position": [1, 2, 3]}},
                "render": {"image_data_url": "data:image/png;base64,x", "submitted_at": "2026-08-07T00:00:00Z"},
            },
        ],
    }


def test_design_delivery_package_includes_proposal_record(tmp_path, monkeypatch) -> None:
    client, project_id = _client(tmp_path, monkeypatch)
    response = client.post(
        f"/api/projects/{project_id}/design-delivery",
        json=_design_delivery_payload(project_id),
    )
    assert response.status_code == 200
    package = response.json()
    assert package["artifact_type"] == "roompilot.web_design_delivery.v1"
    assert package["snapshot_id"] == "snap-1"
    headings = [section["heading"] for section in package["web_report"]["sections"]]
    assert headings[-1] == "六、設計提案 PDF"
    # 尚未產出 PDF 時，成果包仍要回報提案狀態而不是缺欄位。
    assert package["delivery_proposal"]["status"] == "not_generated"
    # 逐房章節：generativeEquipment 與 render 完成度要被帶出。
    room = package["presentation"]["rooms"][0]
    assert room["room_type"] == "living_room"
    assert room["decoration_summary"]["ceiling_and_lighting"] == {"primaryUse": "影音娛樂"}
    assert package["engineering_report"]["completion"]["rendered_room_count"] == 1
    assert package["engineering_report"]["structure_counts"]["walls"] == 2
    # 預算：有目錄價的家具列參考價，其餘與裝潢一律待報價，不得補猜總價。
    budget = package["budget_report"]
    assert budget["known_furniture_reference_subtotal_twd"] == 12900
    statuses = {line["status"] for line in budget["lines"]}
    assert statuses == {"catalog_reference", "pending_quote"}
    # 資安審核：email 這類敏感欄位必須從成果包剔除並列入 redacted_paths。
    assert package["security_review"]["status"] == "passed_with_redactions"
    assert "email" not in package["presentation"]["style_card"]
    assert any("email" in path for path in package["security_review"]["redacted_paths"])


def test_design_delivery_reports_generated_proposal(tmp_path, monkeypatch) -> None:
    client, project_id = _client(tmp_path, monkeypatch)
    store = main.PROJECT_STORE
    store.update_workflow(
        project_id,
        workflow={"delivery_proposal": {"filename": "proposal.pdf", "page_count": 7}},
    )
    response = client.post(
        f"/api/projects/{project_id}/design-delivery",
        json=_design_delivery_payload(project_id),
    )
    assert response.status_code == 200
    proposal = response.json()["delivery_proposal"]
    assert proposal["status"] == "generated"
    assert proposal["download_url"] == f"/api/projects/{project_id}/delivery-proposal/pdf"
    assert "filename" not in proposal


def test_design_delivery_rejects_project_mismatch(tmp_path, monkeypatch) -> None:
    client, project_id = _client(tmp_path, monkeypatch)
    response = client.post(
        f"/api/projects/{project_id}/design-delivery",
        json={"project_id": "someone-else"},
    )
    assert response.status_code == 422
    assert response.json()["detail"]["code"] == "delivery_project_mismatch"


def test_design_delivery_keeps_unpriced_portable_furniture_pending(tmp_path, monkeypatch) -> None:
    client, project_id = _client(tmp_path, monkeypatch)
    payload = _design_delivery_payload(project_id)
    payload["configuration_snapshot"]["furniture"] = [
        {"id": "fixture-sofa-2seat", "room_id": "living", "label": "程序化沙發"},
        {"id": "not-a-catalog-id", "room_id": "living", "label": "自訂家具"},
    ]
    response = client.post(
        f"/api/projects/{project_id}/design-delivery", json=payload
    )
    assert response.status_code == 200
    budget = response.json()["budget_report"]
    furniture_lines = [line for line in budget["lines"] if line["category"] == "furniture"]
    assert budget["known_furniture_reference_subtotal_twd"] == 0
    assert len(furniture_lines) == 2
    assert all(line["amount_twd"] is None for line in furniture_lines)
    assert all(line["status"] == "pending_quote" for line in furniture_lines)


def test_design_delivery_does_not_infer_price_from_unverified_glb_filename(
    tmp_path, monkeypatch
) -> None:
    """實際存檔的家具兩個 id 欄位都不是型錄 id：``furniture_id`` 是引擎擺位 id
    （engine/rules.py 的 room-1-bed-1），``catalog_furniture_id`` 是前端候選槽 id
    （scene_v2.js 的 room-1-bed-double-candidate-1）。此時只剩 model_url 的 GLB
    檔名認得出屋主選了哪一款。

    這是「報價單全是待報價、家具小計 0 元」的真正成因——型錄有價、回查邏輯也對，
    斷的是 join key。
    """
    client, project_id = _client(tmp_path, monkeypatch)
    payload = _design_delivery_payload(project_id)
    payload["configuration_snapshot"]["furniture"] = [
        {
            "furniture_id": "room-1-bed-1",
            "catalog_furniture_id": "room-1-bed-double-candidate-1",
            "room_id": "living",
            "name_zh": "床架",
            "model_url": "https://cdn.example/models/ikea/unverified-bed.glb",
        },
        # 連 GLB 都沒有就真的無從查起，維持待報價、不推估。
        {"furniture_id": "room-1-chair-2", "room_id": "living", "name_zh": "單椅"},
    ]
    response = client.post(
        f"/api/projects/{project_id}/design-delivery", json=payload
    )
    assert response.status_code == 200
    budget = response.json()["budget_report"]
    furniture_lines = [line for line in budget["lines"] if line["category"] == "furniture"]
    assert budget["known_furniture_reference_subtotal_twd"] == 0
    assert all(line["amount_twd"] is None for line in furniture_lines)
    assert budget["pending_quote_count"] >= 2
